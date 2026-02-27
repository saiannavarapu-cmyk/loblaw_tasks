
"""
=============================================================================
  DSG Production Jobs Analyzer
  Analyzes: Box Job | Command | Condition columns from JIL Excel export
=============================================================================
  Usage:
      python analyze_jobs.py
      python analyze_jobs.py --file path/to/your_file.xlsx
      python analyze_jobs.py --file data.xlsx --output results.xlsx
      python analyze_jobs.py --filter-box "mkt-inh-mdw-sdm001_smdwd035.b"
      python analyze_jobs.py --filter-script "CMMDWNewItemsToRB.ksh"
      python analyze_jobs.py --filter-team "mkt"
=============================================================================
"""

import argparse
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Default file path ────────────────────────────────────────────────────────
DEFAULT_FILE = r"C:\Users\sai.annavarapu\Downloads\DSG_Jobs_Analysis\Received_from_client_Other_System_DSG and Oracle(lakrndwpr01)_All-DSG-prod-jobs-Jil.xlsx"


# ════════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ════════════════════════════════════════════════════════════════════════════

def load_data(filepath: str) -> pd.DataFrame:
    """Load the JIL Excel file and return a cleaned DataFrame."""
    print(f"\n Loading: {filepath}")
    df = pd.read_excel(filepath)
    print(f"   ✔  {len(df)} rows | {len(df.columns)} columns loaded")
    return df


# ════════════════════════════════════════════════════════════════════════════
#  BOX JOB ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

def analyze_box_job(df: pd.DataFrame) -> dict:
    """Analyze the 'Box job' column."""
    total          = len(df)
    has_box        = df['Box job'].notna().sum()
    no_box         = df['Box job'].isna().sum()
    unique_boxes   = df['Box job'].nunique()

    box_counts = df['Box job'].value_counts().reset_index()
    box_counts.columns = ['Box Job', 'Job Count']

    # Extract team prefix from box name  e.g. "mkt", "fin", "scs"
    df['_team'] = df['Box job'].dropna().str.split('-').str[0]
    team_summary = df['_team'].value_counts().reset_index()
    team_summary.columns = ['Team', 'Job Count']

    return {
        'total_jobs'    : total,
        'has_box'       : has_box,
        'no_box'        : no_box,
        'unique_boxes'  : unique_boxes,
        'box_counts'    : box_counts,
        'team_summary'  : team_summary,
    }


# ════════════════════════════════════════════════════════════════════════════
#  COMMAND ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

def parse_command(cmd: str) -> dict:
    """Extract script path, name, extension, directory, and arguments."""
    if pd.isna(cmd):
        return {}
    cmd = str(cmd).strip()

    # Script path extraction
    match = re.search(r'(/opt/\S+?\.(?:ksh|sh)|/\S+?\.(?:ksh|sh))', cmd)
    if not match:
        return {'raw_command': cmd, 'script_path': None, 'script_name': None,
                'extension': None, 'directory': None, 'arguments': cmd}

    script_path = match.group(1)
    parts       = script_path.rsplit('/', 1)
    directory   = parts[0] if len(parts) == 2 else ''
    script_name = parts[-1]
    ext_match   = re.search(r'\.(ksh|sh)$', script_name)
    extension   = ext_match.group(1) if ext_match else None

    # Arguments = everything after the script path
    arg_start = cmd.find(script_path) + len(script_path)
    arguments = cmd[arg_start:].strip()

    return {
        'raw_command' : cmd,
        'script_path' : script_path,
        'script_name' : script_name,
        'extension'   : extension,
        'directory'   : directory,
        'arguments'   : arguments if arguments else None,
    }


def analyze_command(df: pd.DataFrame) -> dict:
    """Analyze the 'Command' column."""
    parsed    = df['Command'].apply(parse_command)
    cmd_df    = pd.DataFrame(list(parsed))

    total_cmds    = df['Command'].notna().sum()
    null_cmds     = df['Command'].isna().sum()
    unique_scripts= cmd_df['script_path'].dropna().nunique()

    # Extension breakdown
    ext_counts = cmd_df['extension'].value_counts().reset_index()
    ext_counts.columns = ['Extension', 'Count']

    # Top scripts by usage
    script_counts = cmd_df['script_path'].dropna().value_counts().reset_index()
    script_counts.columns = ['Script Path', 'Usage Count']
    script_counts['Script Name'] = script_counts['Script Path'].apply(lambda x: x.rsplit('/', 1)[-1] if x else x)

    # Directory breakdown
    dir_counts = cmd_df['directory'].dropna().value_counts().reset_index()
    dir_counts.columns = ['Directory', 'Job Count']

    # Full job+command detail
    detail = df[['Job', 'Box job', 'Command']].copy()
    detail['Script Path'] = cmd_df['script_path'].values
    detail['Script Name'] = cmd_df['script_name'].values
    detail['Extension']   = cmd_df['extension'].values
    detail['Directory']   = cmd_df['directory'].values
    detail['Arguments']   = cmd_df['arguments'].values

    return {
        'total_cmds'    : total_cmds,
        'null_cmds'     : null_cmds,
        'unique_scripts': unique_scripts,
        'ext_counts'    : ext_counts,
        'script_counts' : script_counts,
        'dir_counts'    : dir_counts,
        'detail'        : detail,
    }


# ════════════════════════════════════════════════════════════════════════════
#  CONDITION ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

def parse_condition(cond: str) -> dict:
    """Parse an AutoSys condition string into its components."""
    if pd.isna(cond):
        return {'raw': None, 'dep_count': 0, 'deps': [], 'has_time_offset': False,
                'dep_types': [], 'complexity': 'None'}
    cond = str(cond).strip()

    # Extract all s(job_name[,offset]) tokens
    pattern = r's\(([^,)]+)(?:,(\d+\.\d+))?\)'
    matches = re.findall(pattern, cond)

    deps            = []
    has_time_offset = False
    dep_types       = []

    for job_ref, offset in matches:
        job_ref = job_ref.strip()
        dep_type = 'box' if job_ref.endswith('.b') else 'script'
        deps.append({
            'job'        : job_ref,
            'type'       : dep_type,
            'time_offset': offset if offset else None,
        })
        if offset:
            has_time_offset = True
        dep_types.append(dep_type)

    dep_count = len(deps)
    if dep_count == 0:
        complexity = 'None'
    elif dep_count == 1:
        complexity = 'Simple'
    elif dep_count <= 3:
        complexity = 'Medium'
    else:
        complexity = 'Complex'

    return {
        'raw'            : cond,
        'dep_count'      : dep_count,
        'deps'           : deps,
        'has_time_offset': has_time_offset,
        'dep_types'      : dep_types,
        'complexity'     : complexity,
    }


def analyze_condition(df: pd.DataFrame) -> dict:
    """Analyze the 'Condition' column."""
    parsed    = df['Condition'].apply(parse_condition)
    cond_df   = pd.DataFrame(list(parsed))

    has_cond  = df['Condition'].notna().sum()
    no_cond   = df['Condition'].isna().sum()

    # Complexity distribution
    complexity_counts = cond_df['complexity'].value_counts().reset_index()
    complexity_counts.columns = ['Complexity', 'Count']

    # Time-offset conditions
    time_offset_count = cond_df['has_time_offset'].sum()

    # Dependency type breakdown
    box_deps    = sum(1 for types in cond_df['dep_types'] for t in types if t == 'box')
    script_deps = sum(1 for types in cond_df['dep_types'] for t in types if t == 'script')

    # Most referenced dependency jobs
    all_deps = []
    for idx, row in df.iterrows():
        parsed_c = parse_condition(row['Condition'])
        for dep in parsed_c['deps']:
            all_deps.append({
                'Dependent Job'   : row['Job'],
                'Box'             : row['Box job'],
                'Depends On'      : dep['job'],
                'Dep Type'        : dep['type'],
                'Time Offset'     : dep['time_offset'],
            })
    dep_detail = pd.DataFrame(all_deps)

    most_referenced = pd.DataFrame()
    if not dep_detail.empty:
        most_referenced = dep_detail.groupby('Depends On').size().reset_index(name='Referenced By (# jobs)')
        most_referenced = most_referenced.sort_values('Referenced By (# jobs)', ascending=False)

    # Full condition detail per job
    detail = df[['Job', 'Box job', 'Condition']].copy()
    detail['Dep Count']      = cond_df['dep_count'].values
    detail['Complexity']     = cond_df['complexity'].values
    detail['Has Time Offset'] = cond_df['has_time_offset'].values

    return {
        'has_cond'         : has_cond,
        'no_cond'          : no_cond,
        'complexity_counts': complexity_counts,
        'time_offset_count': time_offset_count,
        'box_deps'         : box_deps,
        'script_deps'      : script_deps,
        'dep_detail'       : dep_detail,
        'most_referenced'  : most_referenced,
        'detail'           : detail,
    }


# ════════════════════════════════════════════════════════════════════════════
#  OPTIONAL FILTERS
# ════════════════════════════════════════════════════════════════════════════

def apply_filters(df: pd.DataFrame, args) -> pd.DataFrame:
    original = len(df)
    if args.filter_box:
        df = df[df['Box job'].fillna('').str.contains(args.filter_box, case=False)]
        print(f"   🔍 Filter box='{args.filter_box}': {len(df)}/{original} rows")
    if args.filter_script:
        df = df[df['Command'].fillna('').str.contains(args.filter_script, case=False)]
        print(f"   🔍 Filter script='{args.filter_script}': {len(df)}/{original} rows")
    if args.filter_team:
        pattern = f"^{args.filter_team}-"
        df = df[df['Box job'].fillna('').str.match(pattern, case=False)]
        print(f"   🔍 Filter team='{args.filter_team}': {len(df)}/{original} rows")
    return df


# ════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT WRITER
# ════════════════════════════════════════════════════════════════════════════

# ── Colour palette ───────────────────────────────────────────────────────────
HDR_BOX  = "1F4E79"   # dark blue
HDR_CMD  = "375623"   # dark green
HDR_COND = "843C0C"   # dark orange
HDR_SUM  = "404040"   # dark grey
WHITE    = "FFFFFF"
LIGHT_BLUE   = "DEEAF1"
LIGHT_GREEN  = "E2EFDA"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GREY   = "F2F2F2"


def hdr_style(ws, row, col, value, bg, fg="FFFFFF", width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=fg, name='Arial', size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def data_cell(ws, row, col, value, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name='Arial', size=9)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    thin = Side(style='thin', color='DDDDDD')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)
    return cell


def write_df_to_sheet(ws, df, hdr_color, alt_color, start_row=1):
    """Write a DataFrame to a worksheet with header + alternating rows."""
    for ci, col_name in enumerate(df.columns, 1):
        hdr_style(ws, start_row, ci, col_name, hdr_color, width=max(len(str(col_name)) + 4, 14))
    ws.row_dimensions[start_row].height = 28

    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        fill = alt_color if (ri - start_row) % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = data_cell(ws, ri, ci, val, fill)
            col_letter = get_column_letter(ci)
            # Auto-width hint
            content_len = len(str(val)) if val is not None else 0
            current = ws.column_dimensions[col_letter].width
            ws.column_dimensions[col_letter].width = min(max(current, content_len + 4), 60)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = ws.cell(row=start_row, column=1).coordinate + \
                         ":" + ws.cell(row=start_row, column=len(df.columns)).coordinate


def write_kv(ws, pairs, hdr_color, start_row=1):
    """Write key-value stat pairs as a vertical table."""
    for i, (label, value) in enumerate(pairs):
        r = start_row + i
        hdr_style(ws, r, 1, label, hdr_color, width=35)
        cell = ws.cell(row=r, column=2, value=value)
        cell.font      = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22
    ws.column_dimensions['B'].width = 18


def build_report(
    df: pd.DataFrame,
    box_res: dict,
    cmd_res: dict,
    cond_res: dict,
    output_path: str,
):
    wb = Workbook()

    # ── 1. Summary Dashboard ─────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells('A1:F1')
    title = ws_sum['A1']
    title.value     = "DSG Production Jobs — Analysis Dashboard"
    title.font      = Font(bold=True, size=16, color=WHITE, name='Arial')
    title.fill      = PatternFill("solid", start_color=HDR_SUM)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 36

    # Section headers
    for col, (label, color) in enumerate(
        [("BOX JOB", HDR_BOX), ("COMMAND", HDR_CMD), ("CONDITION", HDR_COND)], 1
    ):
        hdr_style(ws_sum, 3, col * 2 - 1, label, color, width=30)
        hdr_style(ws_sum, 3, col * 2,     "Value", color, width=16)
        ws_sum.merge_cells(
            start_row=3, start_column=col * 2 - 1,
            end_row=3,   end_column=col * 2
        )
        ws_sum.row_dimensions[3].height = 24

    # Box job stats — col A-B
    box_stats = [
        ("Total Jobs",           box_res['total_jobs']),
        ("Jobs in a Box",        box_res['has_box']),
        ("Standalone (No Box)",  box_res['no_box']),
        ("Unique Boxes",         box_res['unique_boxes']),
    ]
    for i, (k, v) in enumerate(box_stats, 4):
        ws_sum.cell(row=i, column=1, value=k).font = Font(name='Arial', size=10)
        ws_sum.cell(row=i, column=2, value=v).font = Font(name='Arial', size=10, bold=True)

    # Command stats — col C-D
    cmd_stats = [
        ("Total Commands",       cmd_res['total_cmds']),
        ("Null Commands",        cmd_res['null_cmds']),
        ("Unique Script Files",  cmd_res['unique_scripts']),
        (".ksh Scripts",         int(cmd_res['ext_counts'][cmd_res['ext_counts']['Extension'] == 'ksh']['Count'].sum()) if 'ksh' in cmd_res['ext_counts']['Extension'].values else 0),
        (".sh Scripts",          int(cmd_res['ext_counts'][cmd_res['ext_counts']['Extension'] == 'sh']['Count'].sum()) if 'sh' in cmd_res['ext_counts']['Extension'].values else 0),
    ]
    for i, (k, v) in enumerate(cmd_stats, 4):
        ws_sum.cell(row=i, column=3, value=k).font = Font(name='Arial', size=10)
        ws_sum.cell(row=i, column=4, value=v).font = Font(name='Arial', size=10, bold=True)

    # Condition stats — col E-F
    cond_stats = [
        ("Jobs with Condition",       cond_res['has_cond']),
        ("Jobs without Condition",    cond_res['no_cond']),
        ("Time-Offset Conditions",    int(cond_res['time_offset_count'])),
        ("Box-Type Dependencies",     cond_res['box_deps']),
        ("Script-Type Dependencies",  cond_res['script_deps']),
    ]
    for i, (k, v) in enumerate(cond_stats, 4):
        ws_sum.cell(row=i, column=5, value=k).font = Font(name='Arial', size=10)
        ws_sum.cell(row=i, column=6, value=v).font = Font(name='Arial', size=10, bold=True)

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_sum.column_dimensions[col].width = 32 if col in ['A', 'C', 'E'] else 14

    # ── 2. Box Job Sheets ────────────────────────────────────────────────────
    ws_box = wb.create_sheet(" Box Job Counts")
    write_df_to_sheet(ws_box, box_res['box_counts'], HDR_BOX, LIGHT_BLUE)

    ws_team = wb.create_sheet(" Team Summary")
    write_df_to_sheet(ws_team, box_res['team_summary'], HDR_BOX, LIGHT_BLUE)

    # ── 3. Command Sheets ────────────────────────────────────────────────────
    ws_scripts = wb.create_sheet(" Script Usage")
    write_df_to_sheet(ws_scripts, cmd_res['script_counts'], HDR_CMD, LIGHT_GREEN)

    ws_dirs = wb.create_sheet(" Directories")
    write_df_to_sheet(ws_dirs, cmd_res['dir_counts'], HDR_CMD, LIGHT_GREEN)

    ws_cmd_detail = wb.create_sheet(" Command Detail")
    write_df_to_sheet(ws_cmd_detail, cmd_res['detail'], HDR_CMD, LIGHT_GREEN)

    # ── 4. Condition Sheets ──────────────────────────────────────────────────
    ws_cond_sum = wb.create_sheet(" Condition Complexity")
    write_df_to_sheet(ws_cond_sum, cond_res['complexity_counts'], HDR_COND, LIGHT_ORANGE)

    ws_cond_detail = wb.create_sheet(" Condition Detail")
    write_df_to_sheet(ws_cond_detail, cond_res['detail'], HDR_COND, LIGHT_ORANGE)

    ws_deps = wb.create_sheet(" All Dependencies")
    if not cond_res['dep_detail'].empty:
        write_df_to_sheet(ws_deps, cond_res['dep_detail'], HDR_COND, LIGHT_ORANGE)

    ws_refs = wb.create_sheet(" Most Referenced")
    if not cond_res['most_referenced'].empty:
        write_df_to_sheet(ws_refs, cond_res['most_referenced'], HDR_COND, LIGHT_ORANGE)

    # ── 5. Full Data ─────────────────────────────────────────────────────────
    ws_all = wb.create_sheet(" Full Data")
    cols = ['Job', 'job_type', 'Box job', 'Command', 'Condition',
            'machine', 'owner', 'start_times', 'days_of_week',
            'Description', 'Support Group']
    available = [c for c in cols if c in df.columns]
    write_df_to_sheet(ws_all, df[available], HDR_SUM, LIGHT_GREY)

    wb.save(output_path)
    print(f"\n  Report saved → {output_path}")


# ════════════════════════════════════════════════════════════════════════════
#  CONSOLE SUMMARY PRINTER
# ════════════════════════════════════════════════════════════════════════════

def print_summary(box_res, cmd_res, cond_res):
    print("\n" + "═" * 60)
    print("    BOX JOB ANALYSIS")
    print("═" * 60)
    print(f"  Total Jobs          : {box_res['total_jobs']}")
    print(f"  Jobs in a Box       : {box_res['has_box']}")
    print(f"  Standalone (no box) : {box_res['no_box']}")
    print(f"  Unique Boxes        : {box_res['unique_boxes']}")
    print("\n  Top 5 Boxes by Job Count:")
    for _, r in box_res['box_counts'].head(5).iterrows():
        print(f"    {r['Box Job']:<55} {r['Job Count']:>3} jobs")

    print("\n" + "═" * 60)
    print("     COMMAND ANALYSIS")
    print("═" * 60)
    print(f"  Total Commands      : {cmd_res['total_cmds']}")
    print(f"  Null Commands       : {cmd_res['null_cmds']}")
    print(f"  Unique Script Files : {cmd_res['unique_scripts']}")
    print("\n  Extension Breakdown:")
    for _, r in cmd_res['ext_counts'].iterrows():
        print(f"    .{r['Extension']:<8} {r['Count']:>4} jobs")
    print("\n  Top 5 Scripts by Usage:")
    for _, r in cmd_res['script_counts'].head(5).iterrows():
        print(f"    {r['Script Name']:<55} {r['Usage Count']:>3} uses")

    print("\n" + "═" * 60)
    print("    CONDITION ANALYSIS")
    print("═" * 60)
    print(f"  Has Condition       : {cond_res['has_cond']}")
    print(f"  No Condition        : {cond_res['no_cond']}")
    print(f"  With Time Offset    : {cond_res['time_offset_count']}")
    print(f"  Box-Type Deps       : {cond_res['box_deps']}")
    print(f"  Script-Type Deps    : {cond_res['script_deps']}")
    print("\n  Complexity Breakdown:")
    for _, r in cond_res['complexity_counts'].iterrows():
        print(f"    {r['Complexity']:<12} {r['Count']:>4} jobs")
    if not cond_res['most_referenced'].empty:
        print("\n  Top 5 Most-Referenced Dependencies:")
        for _, r in cond_res['most_referenced'].head(5).iterrows():
            print(f"    {r['Depends On']:<55} {r['Referenced By (# jobs)']:>3} jobs")
    print("═" * 60)


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Analyze Box Job, Command, and Condition columns from DSG JIL Excel."
    )
    parser.add_argument('--file',          default=DEFAULT_FILE, help='Path to the input .xlsx file')
    parser.add_argument('--output',        default=r"C:\Users\sai.annavarapu\Downloads\DSG_Jobs_Analysis\FINAL_REPORT.xlsx",help='Output Excel report path')
    parser.add_argument('--filter-box',    default=None, help='Filter rows by Box job name (substring match)')
    parser.add_argument('--filter-script', default=None, help='Filter rows by Command script name (substring match)')
    parser.add_argument('--filter-team',   default=None, help='Filter rows by team prefix, e.g. "mkt", "fin"')
    parser.add_argument('--no-excel',      action='store_true', help='Skip Excel report generation')
    args = parser.parse_args()

    # Load
    df = load_data(args.file)

    # Apply filters
    print("\n Applying filters...")
    df = apply_filters(df, args)
    if df.empty:
        print("  No rows match the given filters. Exiting.")
        sys.exit(1)

    # Analyse
    print("\n Analysing columns...")
    box_res  = analyze_box_job(df)
    cmd_res  = analyze_command(df)
    cond_res = analyze_condition(df)

    # Console output
    print_summary(box_res, cmd_res, cond_res)

    # Excel report
    if not args.no_excel:
        build_report(df, box_res, cmd_res, cond_res, args.output)
        print(f"\n  Open '{args.output}' to explore all sheets.")
    else:
        print("\n(Excel report skipped via --no-excel flag)")


if __name__ == "__main__":
    main()