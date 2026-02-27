"""
Shell Script → Job → Box Lineage Analyzer
Produces a multi-sheet Excel report grouping jobs by the script they call.
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

INPUT_FILE  = r"C:\Users\sai.annavarapu\Downloads\shell_Utility\Received_from_client_Other_System_DSG and Oracle(lakrndwpr01)_All-DSG-prod-jobs-Jil.xlsx"
OUTPUT_FILE = r"C:\Users\sai.annavarapu\Downloads\shell_Utility\output.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
C_TITLE      = "1F3864"   # navy   – title bar
C_HDR_SHARED = "1F4E79"   # blue   – shared-script header
C_HDR_SINGLE = "375623"   # green  – single-use header
C_HDR_SUMM   = "404040"   # grey   – summary
C_HDR_DIR    = "4A235A"   # purple – directory view

# Script-group alternating row bands (10 colours cycling)
BAND_COLORS = [
    "D6E4F0","FFF2CC","E2EFDA","FCE4D6","EAD1DC",
    "D9EAD3","CFE2F3","FFE6CC","E8D5B7","D5E8D4",
]
SINGLE_ROW   = "F0F0F0"
WHITE        = "FFFFFF"

thin  = Side(style='thin',   color='AAAAAA')
thick = Side(style='medium', color='888888')
BORDER_THIN  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)


# ════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════════════

def extract_script(cmd):
    if pd.isna(cmd):
        return None
    m = re.search(r'(/\S+?\.(?:ksh|sh))', str(cmd))
    return m.group(1) if m else None


def extract_args(cmd, script_path):
    """Everything after the script path."""
    if pd.isna(cmd) or pd.isna(script_path) or not script_path:
        return None
    s = str(cmd)
    idx = s.find(str(script_path))
    if idx == -1:
        return None
    rest = s[idx + len(script_path):].strip()
    return rest if rest else None


def cell_style(ws, row, col, value="",
               bold=False, font_size=9, font_color="000000",
               bg=None, h_align="left", v_align="center",
               wrap=True, border=BORDER_THIN, width_hint=None, height=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Calibri', size=font_size, bold=bold, color=font_color)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    c.border    = border
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if width_hint:
        col_letter = get_column_letter(col)
        cur = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = min(max(cur, width_hint), 65)
    if height:
        ws.row_dimensions[row].height = height
    return c


def autofit(ws, col, value):
    letter = get_column_letter(col)
    needed = min(len(str(value)) + 4, 65)
    if ws.column_dimensions[letter].width < needed:
        ws.column_dimensions[letter].width = needed


# ════════════════════════════════════════════════════════════════════════════
#  DATA PREPARATION
# ════════════════════════════════════════════════════════════════════════════

def prepare(filepath):
    df = pd.read_excel(filepath)
    df['_script_path'] = df['Command'].apply(extract_script)
    df['_script_name'] = df['_script_path'].apply(
        lambda x: x.rsplit('/', 1)[-1] if pd.notna(x) else None)
    df['_directory']   = df['_script_path'].apply(
        lambda x: x.rsplit('/', 1)[0] if pd.notna(x) else None)
    df['_arguments']   = df.apply(
        lambda r: extract_args(r['Command'], r['_script_path']), axis=1)
    df['_app_area']    = df['_directory'].apply(
        lambda x: x.split('/')[4] if pd.notna(x) and len(x.split('/')) > 4 else None)

    # Script-level summary
    summary = (
        df[df['_script_path'].notna()]
        .groupby(['_script_path', '_script_name', '_directory', '_app_area'])
        .agg(total_jobs=('Job', 'count'),
             unique_boxes=('Box job', 'nunique'),
             jobs=('Job', list),
             boxes=('Box job', list))
        .reset_index()
        .sort_values(['total_jobs', '_script_name'], ascending=[False, True])
    )
    summary['is_shared'] = summary['total_jobs'] > 1

    return df, summary


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 1 – SUMMARY DASHBOARD
# ════════════════════════════════════════════════════════════════════════════

def write_summary_sheet(wb, df, summary):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # ── title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    cell_style(ws, 1, 1,
               value="DSG Production Jobs — Shell Script Lineage Report",
               bold=True, font_size=16, font_color=WHITE,
               bg=C_TITLE, h_align="center", height=38)

    # ── section: KPIs ────────────────────────────────────────────────────────
    kpis = [
        ("Total Jobs in File",           len(df)),
        ("Jobs with a Shell Script",      df['_script_path'].notna().sum()),
        ("Jobs without a Shell Script",   df['_script_path'].isna().sum()),
        ("Unique Shell Scripts",          summary['_script_path'].nunique()),
        ("Scripts Used by > 1 Job",       summary['is_shared'].sum()),
        ("Scripts Used by Only 1 Job",    (~summary['is_shared']).sum()),
        ("Scripts Spanning > 1 Box",      (summary['unique_boxes'] > 1).sum()),
        ("Unique Box Jobs Involved",       df['Box job'].nunique()),
    ]
    cell_style(ws, 3, 1, "KEY METRICS", bold=True, font_size=11,
               font_color=WHITE, bg=C_HDR_SUMM, h_align="center", height=26)
    ws.merge_cells("A3:B3")

    for i, (label, val) in enumerate(kpis, 4):
        cell_style(ws, i, 1, label, bold=False, font_size=10, bg="F5F5F5",
                   width_hint=40, height=20)
        cell_style(ws, i, 2, val,   bold=True,  font_size=11, bg="EBF3FB",
                   h_align="center", width_hint=12)

    # ── section: Top 10 Shared Scripts ───────────────────────────────────────
    r = 3
    headers = ["Script Name", "Full Path", "App Area", "# Jobs", "# Boxes", "Shared?"]
    widths  = [32, 60, 12, 8, 8, 9]
    cell_style(ws, r, 4, "TOP SCRIPTS BY USAGE", bold=True, font_size=11,
               font_color=WHITE, bg=C_HDR_SHARED, h_align="center", height=26)
    ws.merge_cells(f"D3:I3")

    for ci, (h, w) in enumerate(zip(headers, widths), 4):
        cell_style(ws, 4, ci, h, bold=True, font_size=9, font_color=WHITE,
                   bg=C_HDR_SHARED, h_align="center", width_hint=w, height=22)

    for ri, (_, row) in enumerate(summary.head(20).iterrows(), 5):
        bg = "D6E4F0" if ri % 2 == 0 else WHITE
        vals = [
            row['_script_name'], row['_script_path'], row['_app_area'],
            row['total_jobs'], row['unique_boxes'],
            " Shared" if row['is_shared'] else "—"
        ]
        for ci, v in enumerate(vals, 4):
            cell_style(ws, ri, ci, v, font_size=9, bg=bg, height=18)

    # ── section: App-area breakdown ──────────────────────────────────────────
    app_summary = (
        summary.groupby('_app_area')
        .agg(scripts=('_script_path', 'count'),
             total_jobs=('total_jobs', 'sum'),
             shared=('is_shared', 'sum'))
        .reset_index()
        .sort_values('total_jobs', ascending=False)
        .rename(columns={'_app_area': 'App Area', 'scripts': 'Scripts',
                         'total_jobs': 'Total Jobs', 'shared': 'Shared Scripts'})
    )
    base_r = len(kpis) + 6
    cell_style(ws, base_r, 1, "BREAKDOWN BY APP AREA", bold=True, font_size=11,
               font_color=WHITE, bg=C_HDR_DIR, h_align="center", height=26)
    ws.merge_cells(f"A{base_r}:D{base_r}")
    for ci, h in enumerate(app_summary.columns, 1):
        cell_style(ws, base_r + 1, ci, h, bold=True, font_size=9,
                   font_color=WHITE, bg=C_HDR_DIR, h_align="center", height=20)
    for ri, (_, row) in enumerate(app_summary.iterrows(), base_r + 2):
        bg = "EDE7F6" if ri % 2 == 0 else WHITE
        for ci, v in enumerate(row, 1):
            cell_style(ws, ri, ci, v, font_size=9, bg=bg, height=18)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 2 – SHARED SCRIPT LINEAGE (grouped)
# ════════════════════════════════════════════════════════════════════════════

def write_lineage_sheet(wb, df, summary, sheet_name, shared_only=True, hdr_color=C_HDR_SHARED, band=True):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    subset = summary[summary['is_shared'] == shared_only].copy() if shared_only is not None else summary.copy()

    # Column headers
    COLS = ["#", "Script Name", "Full Script Path", "App Area",
            "Job Name", "Box Job", "Machine", "Arguments", "Condition", "Description"]
    WIDTHS = [5, 30, 55, 12, 42, 42, 16, 35, 45, 50]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title_text = ("Shared Scripts — Multiple Jobs Calling Same Script"
                  if shared_only else "Single-Use Scripts — One Job per Script")
    cell_style(ws, 1, 1, title_text, bold=True, font_size=13,
               font_color=WHITE, bg=C_TITLE, h_align="center", height=32)

    # Header row
    for ci, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell_style(ws, 2, ci, h, bold=True, font_size=9, font_color=WHITE,
                   bg=hdr_color, h_align="center", width_hint=w, height=24)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    row_idx   = 3
    band_idx  = 0
    seq_no    = 0

    for _, script_row in subset.iterrows():
        script_path = script_row['_script_path']
        script_name = script_row['_script_name']
        app_area    = script_row['_app_area']
        job_count   = script_row['total_jobs']

        # All jobs for this script
        jobs_df = df[df['_script_path'] == script_path].copy()

        # Pick band colour
        if band:
            bg = BAND_COLORS[band_idx % len(BAND_COLORS)]
            bg_alt = _darken(bg)
        else:
            bg = SINGLE_ROW
            bg_alt = "E8E8E8"

        # ── Script group header row ──────────────────────────────────────────
        seq_no += 1
        HDR_BG = hdr_color
        cell_style(ws, row_idx, 1,  seq_no,      bold=True, font_size=9,
                   font_color=WHITE, bg=HDR_BG, h_align="center", height=22)
        cell_style(ws, row_idx, 2,  script_name, bold=True, font_size=10,
                   font_color=WHITE, bg=HDR_BG, height=22)
        cell_style(ws, row_idx, 3,  script_path, bold=True, font_size=9,
                   font_color=WHITE, bg=HDR_BG, height=22)
        cell_style(ws, row_idx, 4,  app_area,    bold=True, font_size=9,
                   font_color=WHITE, bg=HDR_BG, h_align="center", height=22)
        # Merge remaining cols with job count badge
        badge = f"▶  {job_count} job(s) call this script"
        for ci in range(5, len(COLS) + 1):
            cell_style(ws, row_idx, ci, badge if ci == 5 else "",
                       bold=True, font_size=9, font_color=WHITE, bg=HDR_BG, height=22)
        row_idx += 1

        # ── One row per job ──────────────────────────────────────────────────
        for ji, (_, job) in enumerate(jobs_df.iterrows()):
            row_bg = bg if ji % 2 == 0 else bg_alt
            cell_style(ws, row_idx, 1,  "",                  bg=row_bg, height=18)
            cell_style(ws, row_idx, 2,  script_name,         bg=row_bg, font_size=9)
            cell_style(ws, row_idx, 3,  script_path,         bg=row_bg, font_size=8)
            cell_style(ws, row_idx, 4,  app_area,            bg=row_bg, font_size=9, h_align="center")
            cell_style(ws, row_idx, 5,  job['Job'],          bg=row_bg, font_size=9, bold=True)
            cell_style(ws, row_idx, 6,  job['Box job'],      bg=row_bg, font_size=9)
            cell_style(ws, row_idx, 7,  job['machine'],      bg=row_bg, font_size=9, h_align="center")
            cell_style(ws, row_idx, 8,  job['_arguments'],   bg=row_bg, font_size=8)
            cell_style(ws, row_idx, 9,  job['Condition'],    bg=row_bg, font_size=8)
            cell_style(ws, row_idx, 10, job['Description'],  bg=row_bg, font_size=8)
            row_idx += 1

        band_idx += 1

    ws.row_dimensions[2].height = 26
    return ws


def _darken(hex_color, factor=0.92):
    """Slightly darken a hex color for alternating rows."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r * factor); g = int(g * factor); b = int(b * factor)
    return f"{r:02X}{g:02X}{b:02X}"


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 3 – DIRECTORY LINEAGE VIEW
# ════════════════════════════════════════════════════════════════════════════

def write_directory_sheet(wb, df, summary):
    ws = wb.create_sheet("Directory")
    ws.sheet_view.showGridLines = False

    COLS   = ["Directory", "Script Name", "# Jobs", "# Boxes", "Jobs List", "Boxes List"]
    WIDTHS = [45, 36, 8, 8, 80, 80]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    cell_style(ws, 1, 1, "Script Lineage — Grouped by Directory",
               bold=True, font_size=13, font_color=WHITE,
               bg=C_TITLE, h_align="center", height=32)

    for ci, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell_style(ws, 2, ci, h, bold=True, font_size=9,
                   font_color=WHITE, bg=C_HDR_DIR, h_align="center",
                   width_hint=w, height=24)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    row_idx  = 3
    dir_idx  = 0
    dirs = summary.groupby('_directory')

    purple_bands = ["EDE7F6","E1D5F5","D5C4F3","C9B4F0","BDA3ED"]

    for dir_path, grp in sorted(dirs, key=lambda x: x[0] or ""):
        grp = grp.sort_values('total_jobs', ascending=False)

        # Directory header
        dir_bg = purple_bands[dir_idx % len(purple_bands)]
        ws.merge_cells(f"A{row_idx}:{get_column_letter(len(COLS))}{row_idx}")
        cell_style(ws, row_idx, 1, f"  {dir_path}",
                   bold=True, font_size=10, bg=C_HDR_DIR,
                   font_color=WHITE, height=22)
        row_idx += 1

        for ji, (_, srow) in enumerate(grp.iterrows()):
            bg = dir_bg if ji % 2 == 0 else _darken(dir_bg)
            jobs_str  = "\n".join(sorted(set(srow['jobs'])))
            boxes_str = "\n".join(sorted(set(str(b) for b in srow['boxes'] if pd.notna(b))))
            cell_style(ws, row_idx, 1, dir_path,          bg=bg, font_size=8)
            cell_style(ws, row_idx, 2, srow['_script_name'], bg=bg, font_size=9, bold=True)
            cell_style(ws, row_idx, 3, srow['total_jobs'],   bg=bg, font_size=9, h_align="center")
            cell_style(ws, row_idx, 4, srow['unique_boxes'],  bg=bg, font_size=9, h_align="center")
            cell_style(ws, row_idx, 5, jobs_str,           bg=bg, font_size=8)
            cell_style(ws, row_idx, 6, boxes_str,          bg=bg, font_size=8)
            lines = max(jobs_str.count('\n'), boxes_str.count('\n')) + 1
            ws.row_dimensions[row_idx].height = max(18, lines * 14)
            row_idx += 1

        dir_idx += 1


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 4 – BOX-CENTRIC LINEAGE
# ════════════════════════════════════════════════════════════════════════════

def write_box_centric_sheet(wb, df):
    ws = wb.create_sheet("Box -> Script Lineage")
    ws.sheet_view.showGridLines = False

    COLS   = ["Box Job", "Job Name", "Script Name", "Full Script Path",
              "App Area", "Arguments", "Condition"]
    WIDTHS = [42, 42, 30, 55, 12, 35, 45]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    cell_style(ws, 1, 1, "Box-Centric View — Each Box → Jobs → Scripts Called",
               bold=True, font_size=13, font_color=WHITE,
               bg=C_TITLE, h_align="center", height=32)

    for ci, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell_style(ws, 2, ci, h, bold=True, font_size=9,
                   font_color=WHITE, bg=C_HDR_SHARED, h_align="center",
                   width_hint=w, height=24)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    row_idx = 3
    band_idx = 0

    valid = df[df['_script_path'].notna()].copy()
    valid['Box job'] = valid['Box job'].fillna("(Standalone — No Box)")

    for box_name, grp in sorted(valid.groupby('Box job'), key=lambda x: x[0]):
        bg = BAND_COLORS[band_idx % len(BAND_COLORS)]

        # Box header
        ws.merge_cells(f"A{row_idx}:{get_column_letter(len(COLS))}{row_idx}")
        cell_style(ws, row_idx, 1,
                   f"  {box_name}   [{len(grp)} job(s)]",
                   bold=True, font_size=10, font_color=WHITE,
                   bg=C_HDR_SHARED, height=22)
        row_idx += 1

        for ji, (_, job) in enumerate(grp.iterrows()):
            row_bg = bg if ji % 2 == 0 else _darken(bg)
            cell_style(ws, row_idx, 1, box_name,              bg=row_bg, font_size=9)
            cell_style(ws, row_idx, 2, job['Job'],            bg=row_bg, font_size=9, bold=True)
            cell_style(ws, row_idx, 3, job['_script_name'],   bg=row_bg, font_size=9)
            cell_style(ws, row_idx, 4, job['_script_path'],   bg=row_bg, font_size=8)
            cell_style(ws, row_idx, 5, job['_app_area'],      bg=row_bg, font_size=9, h_align="center")
            cell_style(ws, row_idx, 6, job['_arguments'],     bg=row_bg, font_size=8)
            cell_style(ws, row_idx, 7, job['Condition'],      bg=row_bg, font_size=8)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        band_idx += 1


# ════════════════════════════════════════════════════════════════════════════
#  SHEET 5 – FLAT DETAIL (filterable)
# ════════════════════════════════════════════════════════════════════════════

def write_flat_sheet(wb, df):
    ws = wb.create_sheet(" Full Flat Detail")
    ws.sheet_view.showGridLines = False

    cols_map = {
        'Job'          : 'Job Name',
        'Box job'      : 'Box Job',
        '_script_name' : 'Script Name',
        '_script_path' : 'Full Script Path',
        '_directory'   : 'Directory',
        '_app_area'    : 'App Area',
        '_arguments'   : 'Arguments',
        'Condition'    : 'Condition',
        'machine'      : 'Machine',
        'Description'  : 'Description',
        'start_times'  : 'Start Times',
        'days_of_week' : 'Days of Week',
    }
    available = {k: v for k, v in cols_map.items() if k in df.columns}
    out = df[list(available.keys())].copy().rename(columns=available)
    out['Shared Script?'] = df['_script_path'].map(
        df.groupby('_script_path')['Job'].transform('count') > 1
    ).map({True: ' Yes', False: '—'})
    out = out.sort_values(['Script Name', 'Box Job', 'Job Name'], na_position='last')

    ws.merge_cells(f"A1:{get_column_letter(len(out.columns))}1")
    cell_style(ws, 1, 1, "Full Flat Detail — All Jobs with Script Lineage",
               bold=True, font_size=13, font_color=WHITE,
               bg=C_TITLE, h_align="center", height=32)

    for ci, h in enumerate(out.columns, 1):
        cell_style(ws, 2, ci, h, bold=True, font_size=9,
                   font_color=WHITE, bg=C_HDR_SUMM,
                   h_align="center", width_hint=max(len(h)+4, 14), height=24)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(out.columns))}2"

    for ri, (_, row) in enumerate(out.iterrows(), 3):
        bg = "F5F5F5" if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row, 1):
            v = val if pd.notna(val) else ""
            c = cell_style(ws, ri, ci, v, bg=bg, font_size=9, height=18)
            autofit(ws, ci, v)


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    print(f"  Loading {INPUT_FILE} …")
    df, summary = prepare(INPUT_FILE)
    print(f"   ✔  {len(df)} jobs | {summary['_script_path'].nunique()} unique scripts")

    wb = Workbook()

    print("  Writing Summary …")
    write_summary_sheet(wb, df, summary)

    print("  Writing Shared Script Lineage …")
    write_lineage_sheet(wb, df, summary,
                        sheet_name="Shared Script Lineage",
                        shared_only=True,
                        hdr_color=C_HDR_SHARED)

    print("  Writing Single-Use Script Lineage …")
    write_lineage_sheet(wb, df, summary,
                        sheet_name="Single-Use Scripts",
                        shared_only=False,
                        hdr_color=C_HDR_SINGLE,
                        band=False)

    print("  Writing Directory View …")
    write_directory_sheet(wb, df, summary)

    print("  Writing Box-Centric View …")
    write_box_centric_sheet(wb, df)

    print("  Writing Full Flat Detail …")
    write_flat_sheet(wb, df)

    wb.save(OUTPUT_FILE)
    print(f"\n  Saved → {OUTPUT_FILE}")
    print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
